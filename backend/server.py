from fastapi import FastAPI, APIRouter, HTTPException, Depends
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# JWT Config
JWT_SECRET = os.environ.get('JWT_SECRET', 'boxtrack-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer()

# ==================== AUTH MODELS ====================

class UserCreate(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)
    password: str = Field(..., min_length=6)

class UserLogin(BaseModel):
    username: str
    password: str

class UserResponse(BaseModel):
    id: str
    username: str
    created_at: str

class TokenResponse(BaseModel):
    token: str
    user: UserResponse

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, username: str) -> str:
    payload = {
        "user_id": user_id,
        "username": username,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("user_id")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user: UserCreate):
    # Check if username exists
    existing = await db.users.find_one({"username": user.username.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Username already taken")
    
    # Create user
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "username": user.username.lower(),
        "password_hash": hash_password(user.password),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    # Generate token
    token = create_token(user_id, user.username.lower())
    
    return TokenResponse(
        token=token,
        user=UserResponse(id=user_id, username=user.username.lower(), created_at=user_doc["created_at"])
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"username": credentials.username.lower()}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    if not verify_password(credentials.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    token = create_token(user["id"], user["username"])
    
    return TokenResponse(
        token=token,
        user=UserResponse(id=user["id"], username=user["username"], created_at=user["created_at"])
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user["user_id"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return UserResponse(id=user["id"], username=user["username"], created_at=user["created_at"])

# ==================== BOX MODELS ====================

class BoxType(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    quantity: int = 0
    cost: float = 0.0
    min_threshold: int = 10
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class BoxTypeCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    quantity: int = Field(default=0, ge=0)
    cost: float = Field(default=0.0, ge=0)
    min_threshold: int = Field(default=10, ge=0)

class BoxTypeUpdate(BaseModel):
    name: Optional[str] = None
    quantity: Optional[int] = None
    cost: Optional[float] = None
    min_threshold: Optional[int] = None

class UsageRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    box_type_id: str
    box_name: str
    quantity_used: int
    date: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class UsageRecordCreate(BaseModel):
    box_type_id: str
    quantity_used: int
    date: Optional[str] = None

class BoxWithPrediction(BaseModel):
    id: str
    name: str
    quantity: int
    cost: float
    min_threshold: int
    created_at: str
    avg_daily_usage: float
    days_until_empty: Optional[int]
    days_until_reorder: Optional[int]
    prediction_status: str  # "safe", "warning", "critical"

class DashboardStats(BaseModel):
    total_box_types: int
    total_inventory: int
    total_value: float
    low_stock_count: int
    low_stock_boxes: List[dict]
    boxes_needing_reorder_soon: List[dict]

# ==================== PREDICTION HELPERS ====================

async def calculate_box_predictions(box: dict) -> dict:
    """Calculate usage predictions for a box based on historical data"""
    box_id = box["id"]
    
    # Get usage records from last 30 days
    cutoff_date = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    usage_records = await db.usage.find(
        {"box_type_id": box_id, "date": {"$gte": cutoff_date}},
        {"_id": 0}
    ).to_list(1000)
    
    total_used = sum(r["quantity_used"] for r in usage_records)
    
    # Calculate average daily usage
    if usage_records:
        # Get unique days with usage
        unique_dates = set(r["date"] for r in usage_records)
        days_with_data = len(unique_dates)
        # Use 30 days or actual days of data, whichever gives more accurate average
        days_for_calc = max(days_with_data, 7)  # At least 7 days to smooth out
        avg_daily = total_used / days_for_calc
    else:
        avg_daily = 0
    
    # Calculate predictions
    quantity = box["quantity"]
    min_threshold = box["min_threshold"]
    
    if avg_daily > 0:
        days_until_empty = int(quantity / avg_daily) if quantity > 0 else 0
        days_until_reorder = int((quantity - min_threshold) / avg_daily) if quantity > min_threshold else 0
    else:
        days_until_empty = None  # No usage data
        days_until_reorder = None
    
    # Determine status
    if quantity <= min_threshold:
        status = "critical"
    elif days_until_reorder is not None and days_until_reorder <= 7:
        status = "warning"
    elif days_until_empty is not None and days_until_empty <= 14:
        status = "warning"
    else:
        status = "safe"
    
    return {
        **box,
        "avg_daily_usage": round(avg_daily, 2),
        "days_until_empty": days_until_empty,
        "days_until_reorder": days_until_reorder,
        "prediction_status": status
    }

# ==================== BOX ENDPOINTS ====================

@api_router.get("/")
async def root():
    return {"message": "Box Inventory API"}

@api_router.get("/boxes", response_model=List[BoxWithPrediction])
async def get_boxes(current_user: dict = Depends(get_current_user)):
    boxes = await db.boxes.find({}, {"_id": 0}).to_list(1000)
    # Add predictions to each box
    boxes_with_predictions = []
    for box in boxes:
        box_with_pred = await calculate_box_predictions(box)
        boxes_with_predictions.append(box_with_pred)
    return boxes_with_predictions

@api_router.post("/boxes", response_model=BoxType)
async def create_box(box: BoxTypeCreate, current_user: dict = Depends(get_current_user)):
    box_obj = BoxType(**box.model_dump())
    doc = box_obj.model_dump()
    await db.boxes.insert_one(doc)
    return box_obj

@api_router.get("/boxes/{box_id}", response_model=BoxWithPrediction)
async def get_box(box_id: str, current_user: dict = Depends(get_current_user)):
    box = await db.boxes.find_one({"id": box_id}, {"_id": 0})
    if not box:
        raise HTTPException(status_code=404, detail="Box not found")
    return await calculate_box_predictions(box)

@api_router.put("/boxes/{box_id}", response_model=BoxType)
async def update_box(box_id: str, box_update: BoxTypeUpdate, current_user: dict = Depends(get_current_user)):
    existing = await db.boxes.find_one({"id": box_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Box not found")
    
    update_data = {k: v for k, v in box_update.model_dump().items() if v is not None}
    if update_data:
        await db.boxes.update_one({"id": box_id}, {"$set": update_data})
    
    updated = await db.boxes.find_one({"id": box_id}, {"_id": 0})
    return updated

@api_router.delete("/boxes/{box_id}")
async def delete_box(box_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.boxes.delete_one({"id": box_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Box not found")
    await db.usage.delete_many({"box_type_id": box_id})
    return {"message": "Box deleted successfully"}

# ==================== USAGE ENDPOINTS ====================

@api_router.post("/usage", response_model=UsageRecord)
async def record_usage(usage: UsageRecordCreate, current_user: dict = Depends(get_current_user)):
    box = await db.boxes.find_one({"id": usage.box_type_id}, {"_id": 0})
    if not box:
        raise HTTPException(status_code=404, detail="Box not found")
    
    if box["quantity"] < usage.quantity_used:
        raise HTTPException(status_code=400, detail=f"Not enough stock. Available: {box['quantity']}")
    
    new_quantity = box["quantity"] - usage.quantity_used
    await db.boxes.update_one({"id": usage.box_type_id}, {"$set": {"quantity": new_quantity}})
    
    use_date = usage.date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    usage_obj = UsageRecord(
        box_type_id=usage.box_type_id,
        box_name=box["name"],
        quantity_used=usage.quantity_used,
        date=use_date
    )
    doc = usage_obj.model_dump()
    await db.usage.insert_one(doc)
    
    return usage_obj

@api_router.get("/usage", response_model=List[UsageRecord])
async def get_usage(days: int = 30, current_user: dict = Depends(get_current_user)):
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    records = await db.usage.find(
        {"created_at": {"$gte": cutoff}},
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    return records

@api_router.get("/usage/trends")
async def get_usage_trends(days: int = 14, current_user: dict = Depends(get_current_user)):
    cutoff_date = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
    
    records = await db.usage.find(
        {"date": {"$gte": cutoff_date}},
        {"_id": 0}
    ).to_list(1000)
    
    daily_totals = {}
    for r in records:
        date = r["date"]
        if date not in daily_totals:
            daily_totals[date] = 0
        daily_totals[date] += r["quantity_used"]
    
    result = []
    current = datetime.now(timezone.utc)
    for i in range(days):
        date = (current - timedelta(days=i)).strftime("%Y-%m-%d")
        result.append({
            "date": date,
            "total_used": daily_totals.get(date, 0)
        })
    
    result.reverse()
    return result

# ==================== DASHBOARD ENDPOINTS ====================

@api_router.get("/stats", response_model=DashboardStats)
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    boxes = await db.boxes.find({}, {"_id": 0}).to_list(1000)
    
    total_box_types = len(boxes)
    total_inventory = sum(b["quantity"] for b in boxes)
    total_value = sum(b["quantity"] * b["cost"] for b in boxes)
    
    low_stock_boxes = []
    boxes_needing_reorder = []
    
    for box in boxes:
        pred = await calculate_box_predictions(box)
        
        if box["quantity"] <= box["min_threshold"]:
            low_stock_boxes.append({
                "id": box["id"],
                "name": box["name"],
                "quantity": box["quantity"],
                "min_threshold": box["min_threshold"]
            })
        
        # Boxes that will need reorder within 7 days
        if pred["days_until_reorder"] is not None and pred["days_until_reorder"] <= 7 and pred["days_until_reorder"] > 0:
            boxes_needing_reorder.append({
                "id": box["id"],
                "name": box["name"],
                "quantity": box["quantity"],
                "days_until_reorder": pred["days_until_reorder"],
                "avg_daily_usage": pred["avg_daily_usage"]
            })
    
    return DashboardStats(
        total_box_types=total_box_types,
        total_inventory=total_inventory,
        total_value=round(total_value, 2),
        low_stock_count=len(low_stock_boxes),
        low_stock_boxes=low_stock_boxes,
        boxes_needing_reorder_soon=boxes_needing_reorder
    )

# ==================== EXPORT ENDPOINT ====================

@api_router.get("/export/inventory")
async def export_inventory_excel(current_user: dict = Depends(get_current_user)):
    """Export inventory data to Excel file"""
    boxes = await db.boxes.find({}, {"_id": 0}).to_list(1000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Name", "Quantity", "Cost/Unit ($)", "Total Value ($)", "Min Threshold", "Status", "Avg Daily Usage", "Days Until Empty", "Days Until Reorder"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_num, box in enumerate(boxes, 2):
        pred = await calculate_box_predictions(box)
        
        # Determine status text
        if box["quantity"] == 0:
            status = "Out of Stock"
        elif box["quantity"] <= box["min_threshold"]:
            status = "Low Stock"
        else:
            status = "In Stock"
        
        row_data = [
            box["name"],
            box["quantity"],
            box["cost"],
            round(box["quantity"] * box["cost"], 2),
            box["min_threshold"],
            status,
            pred["avg_daily_usage"],
            pred["days_until_empty"] if pred["days_until_empty"] is not None else "N/A",
            pred["days_until_reorder"] if pred["days_until_reorder"] is not None else "N/A"
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col in [2, 3, 4, 5, 7]:  # Number columns
                cell.alignment = Alignment(horizontal="right")
    
    # Adjust column widths
    column_widths = [25, 12, 15, 15, 15, 12, 18, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with date
    filename = f"inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
